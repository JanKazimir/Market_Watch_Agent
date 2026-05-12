## This script uses the report found in /data/pdfs/reports to update the excel file "url_status" column.

import json
import datetime
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent.parent
today = datetime.date.today().isoformat()

EXCEL_PATH = BASE_DIR / "data" / "list_of_banks_ref_new.xlsx"
REPORT_PATH = BASE_DIR / "data" / "pdfs" / "reports" / f"{today}_pdf_diff_report.json"


def update_pdf_status(report_path: Path = REPORT_PATH, excel_path: Path = EXCEL_PATH):
    """Read the PDF report and write url_status back into the Excel 'Links' sheet."""
    with open(report_path, encoding="utf-8") as f:
        report = json.load(f)

    # Build a url -> status string mapping from the report
    status_map = {}
    for entry in report["entries"]:
        url = entry["url"].strip()
        entry_status = entry.get("status", "")
        detail = entry.get("detail", "")

        if entry_status == "updated":
            status_map[url] = f"works: (updated {report['date']})"
        else:
            status_map[url] = f"error: {detail}"

    if not status_map:
        print("No entries in report to update.")
        return

    # Update Excel in place using openpyxl to preserve other sheets/formatting
    wb = load_workbook(excel_path)
    ws = wb["Links"]

    # Find column indices (1-based) from header row
    headers = {cell.value: cell.column for cell in ws[1]}
    url_col = headers.get("url")
    status_col = headers.get("url_status")

    if not url_col:
        print("ERROR: 'url' column not found in Links sheet.")
        return
    if not status_col:
        print("ERROR: 'url_status' column not found in Links sheet.")
        return

    updated_count = 0
    for row in range(2, ws.max_row + 1):
        cell_url = ws.cell(row=row, column=url_col).value
        if cell_url and str(cell_url).strip() in status_map:
            ws.cell(row=row, column=status_col).value = status_map[str(cell_url).strip()]
            updated_count += 1

    wb.save(excel_path)
    print(f"Updated {updated_count} row(s) in {excel_path.name}")


if __name__ == "__main__":
    update_pdf_status()
